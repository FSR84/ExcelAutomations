from openpyxl.utils.cell import get_column_letter, column_index_from_string
from openpyxl.styles import Font, PatternFill, Border, Side
import openpyxl
import os


# load workbook and define target worksheet by name
curr_dir = os.getcwd() + "\\ExcelAutomations\\Excel Files\\"
excel3 = "all-shifts.xlsx"

wb = openpyxl.load_workbook(curr_dir + excel3)
ws = wb["Shifts"]  # wb.active  # (ActiveSheet in VBA)


# CLEANUP: delete first row if cell B1 is empty
if ws["B1"].value == None:
    ws.delete_rows(1, 1)


# CLEANUP: delete added columns
if ws.max_column > 7:
    print("Columns " + get_column_letter(8) + ":" + get_column_letter(ws.max_column) + " have been deleted.")
    ws.delete_cols(8, ws.max_column - 7) # ws.delete_cols(column_index_from_string("H"), ws.max_column - 7)


# insert 1st row if cell B1 is not empty
if ws["B1"].value != None:
    ws.insert_rows(1, 1)


# add missing column name
if ws["A2"].value == None:
    ws["A2"] = "Index"


# add formulas to cells in the top row; change cells to bold (ws["F1:G1"].font does not work)
ws["F1"].value = "=Subtotal(9,F3:F100)"
ws["G1"].value = "=Subtotal(9,G3:G100)"
ws["F1"].font = Font(bold=True)
ws["G1"].font = Font(bold=True)


# insert formulas to a new column
for cell in ws[get_column_letter(ws.max_column + 1)]:  # first empty column
    if cell.row == 2:
        cell.value = "SUMIFS"
    elif cell.row > 2:
        cell.value = "=SUMIFS(G:G,D:D,D" + str(cell.row) + ")"
for cell in ws[get_column_letter(ws.max_column + 1)]:  # first empty column (again)
    if cell.row == 2:
        cell.value = "INDEX/MATCH"
    elif cell.row > 2:
        cell.value = "=INDEX(E:E,MATCH(D" + str(cell.row) + ",D:D,0))"
# for cell in ws[get_column_letter(ws.max_column + 1)]:  # first empty column (again)
#     if cell.row == 2:
#         cell.value = "LOOKUP 2,1" # gives last result based on 1 or more conditions; should not be used in a column if it has 2 or more conditions due to refreshing speed
#     elif cell.row > 2:
#         cell.value = "=LOOKUP(2,1/((D:D=D" + str(cell.row) + ")*(B:B=B" + str(cell.row) + ")),E:E)"


# change font size and apply bold and interior color to an entire row (each cell separately, needs a separate loop for each row); OpenPyXL column loop ends at the last non-empty column, even if there are empty cells in between
for cell in ws["2:2"]:
    cell.font = Font(size=12, color="403151", bold=True)
    cell.fill = PatternFill(fgColor="E4DFEC", fill_type="solid")


# define current region
lastrow = ws.max_row
lastcol = ws.max_column
lastcolletter = get_column_letter(lastcol)

print(str(lastrow) + " is the last row number.")
print(str(lastcol) + " is the last column number.")
print(lastcolletter + str(lastrow) + " is the last cell in the current region.")


# remove all borders; from row 3 onwards remove interior color (fill)
no_fill = PatternFill(fill_type=None)
side = Side(border_style=None)
no_border = Border(left=side, right=side, top=side, bottom=side)

for row in ws:
    for cell in row:
        cell.border = no_border
        if cell.row > 2:
            cell.fill = no_fill


# add borders to table (limited by current region, ignoring first row)
side = Side(border_style="thin")
border = Border(left=side, right=side, top=side, bottom=side)

for row in ws.iter_rows(min_row=2, max_row=lastrow, min_col=1, max_col=lastcol):
    for cell in row:
        cell.border = border


# save workbook
wb.save(curr_dir + excel3)


# open Excel instance with the workbook
os.chdir(curr_dir)
os.system(excel3)

